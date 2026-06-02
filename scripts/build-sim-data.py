"""
build-sim-data.py — Convert 질문 취합.xlsx to sim-questions.json
Usage: python3 scripts/build-sim-data.py

Reads: C:/Users/iceam/Downloads/질문 취합.xlsx  (8 sheets: 4 days × 유튜브/생방송)
Writes: src/data/sim-questions.json

Anonymisation: Korean names in parens inside 내용 column are replaced
               per-row with (시민A), (시민B), ... — mappings are row-local,
               not global, so names cannot be linked across rows.
Expert names in 답변자 column are kept (public figures).

두 가지 시트 스키마:
  유튜브 시트: 순번 | 지역 | 분임 | 답변자 | 내용 | 예정시간 | ...
  생방송 시트: (지역) | 분임 | 답변자 | 내용 | 예정시간 | ...  (헤더 행 없거나 위치 다름)
"""

import sys
import re
import json
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not found. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

XLSX_PATH = Path("C:/Users/iceam/Downloads/질문 취합.xlsx")
OUT_PATH = Path(__file__).parent.parent / "src" / "data" / "sim-questions.json"

# Korean full-name pattern: 2-4 Hangul chars inside parentheses, optional leading space.
# e.g. (이춘우), ( 박규환), (홍길동)
# We anonymise ALL Korean names in parens — experts are in 답변자 column, not 내용.
NAME_RE = re.compile(r"\(\s*([가-힣]{2,4})\s*\)")

CITIZEN_LABELS = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")


def anonymise_row(text: str) -> str:
    """Replace all (Korean-name) occurrences in a single cell value with
    (시민A), (시민B), ... — mapping is local to this call (per-row)."""
    seen: dict[str, str] = {}
    counter = 0

    def replace(m: re.Match) -> str:
        nonlocal counter
        name = m.group(1)
        if name not in seen:
            if counter < 26:
                label = CITIZEN_LABELS[counter]
            else:
                label = str(counter + 1)
            seen[name] = f"시민{label}"
            counter += 1
        return f"({seen[name]})"

    return NAME_RE.sub(replace, text)


def parse_sheet_name(name: str):
    """Return (day: int, channel: str) or None."""
    m = re.search(r"(\d)일차[_ ]*(유튜브|생방송)", name)
    if m:
        return int(m.group(1)), m.group(2)
    return None


def get_row_values(row) -> list[str]:
    return [str(c).strip() if c is not None else "" for c in row]


def process_youtube_sheet(ws, day: int, channel: str, id_offset: int) -> list[dict]:
    """유튜브 시트: 순번|지역|분임|답변자|내용|예정시간 schema"""
    header_row_num = None
    headers: list[str] = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = get_row_values(row)
        if "순번" in vals and "내용" in vals:
            header_row_num = i
            headers = vals
            break

    if not headers:
        # Fallback: treat first row as header if it has recognisable columns
        first_row = get_row_values(next(ws.iter_rows(values_only=True)))
        if "내용" in first_row:
            header_row_num = 1
            headers = first_row
        else:
            print(f"  WARNING (유튜브): header not found", file=sys.stderr)
            return []

    def ci(name: str):
        try:
            return headers.index(name)
        except ValueError:
            return None

    ci_region  = ci("지역")
    ci_group   = ci("분임")
    ci_expert  = ci("답변자")
    ci_content = ci("내용")
    ci_time    = ci("예정시간")

    if ci_content is None:
        print(f"  WARNING (유튜브): no 내용 column", file=sys.stderr)
        return []

    rows = []
    local_id = id_offset
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        content = row[ci_content] if ci_content is not None else None
        if content is None or str(content).strip() == "":
            continue

        content_str = str(content).strip()
        anon_content = anonymise_row(content_str)

        # Expert column may also have names in col G+ (tally columns)
        expert_val = ""
        if ci_expert is not None and row[ci_expert] is not None:
            expert_val = str(row[ci_expert]).strip()

        rows.append({
            "id": local_id,
            "day": day,
            "channel": channel,
            "region": str(row[ci_region]).strip() if ci_region is not None and row[ci_region] is not None else "",
            "group": str(row[ci_group]).strip() if ci_group is not None and row[ci_group] is not None else "",
            "expert": expert_val,
            "content": anon_content,
            "scheduled_time": str(row[ci_time]).strip() if ci_time is not None and row[ci_time] is not None else "",
            "status": "pending",
        })
        local_id += 1

    return rows


def process_live_sheet(ws, day: int, channel: str, id_offset: int) -> list[dict]:
    """생방송 시트: flexible schema — detect columns by position or header."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    # Try to find a header row
    header_row_num = 0
    headers: list[str] = []
    for i, row in enumerate(all_rows):
        vals = get_row_values(row)
        if "내용" in vals:
            header_row_num = i
            headers = vals
            break

    if not headers:
        # 생방송 시트 1일차: no "내용" column header detected.
        # Check if row 1 has 지역/분임 positionally
        first = get_row_values(all_rows[0])
        # Infer: col0=지역, col1=분임, col2=답변자 or 구분, col3=내용
        # This is a best-effort guess when header is absent
        print(f"  INFO (생방송 day={day}): no 내용 header, attempting positional parse (col3)", file=sys.stderr)
        ci_region, ci_group, ci_expert, ci_content, ci_time = 0, 1, 2, 3, 4
        data_rows = all_rows[1:]  # skip first descriptive row
    else:
        def ci(name: str):
            try:
                return headers.index(name)
            except ValueError:
                return None

        ci_region  = ci("지역") if ci("지역") is not None else 0
        ci_group   = ci("분임") if ci("분임") is not None else 1
        ci_expert  = ci("답변자")
        ci_content = ci("내용")
        ci_time    = ci("예정시간")
        data_rows = all_rows[header_row_num + 1:]

    rows = []
    local_id = id_offset
    for row in data_rows:
        content = row[ci_content] if ci_content is not None and len(row) > ci_content else None
        if content is None or str(content).strip() == "":
            continue

        content_str = str(content).strip()
        anon_content = anonymise_row(content_str)

        expert_val = ""
        if ci_expert is not None and len(row) > ci_expert and row[ci_expert] is not None:
            expert_val = str(row[ci_expert]).strip()

        region_val = str(row[ci_region]).strip() if len(row) > ci_region and row[ci_region] is not None else ""
        group_val  = str(row[ci_group]).strip()  if len(row) > ci_group  and row[ci_group]  is not None else ""
        time_val   = str(row[ci_time]).strip()   if ci_time is not None and len(row) > ci_time and row[ci_time] is not None else ""

        rows.append({
            "id": local_id,
            "day": day,
            "channel": channel,
            "region": region_val,
            "group": group_val,
            "expert": expert_val,
            "content": anon_content,
            "scheduled_time": time_val,
            "status": "pending",
        })
        local_id += 1

    return rows


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx not found at {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    all_rows: list[dict] = []
    id_counter = 1

    for sheet_name in wb.sheetnames:
        parsed = parse_sheet_name(sheet_name)
        if parsed is None:
            print(f"  SKIP sheet '{sheet_name}' — cannot parse day/channel", file=sys.stderr)
            continue

        day, channel = parsed
        ws = wb[sheet_name]

        if channel == "유튜브":
            rows = process_youtube_sheet(ws, day, channel, id_counter)
        else:
            rows = process_live_sheet(ws, day, channel, id_counter)

        id_counter += len(rows)
        all_rows.extend(rows)
        print(f"  Sheet '{sheet_name}' → {len(rows)} rows (day={day}, channel={channel})")

    print(f"\nTotal rows after filtering: {len(all_rows)}")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(all_rows, f, ensure_ascii=False, indent=2)
    print(f"Written to: {OUT_PATH}")

    # Show 3 sample rows
    print("\nSample (3 rows with actual question content):")
    question_rows = [r for r in all_rows if "질문" in r.get("content", "") or "?" in r.get("content", "") or "궁금" in r.get("content", "")]
    for row in question_rows[:3]:
        print(json.dumps(row, ensure_ascii=False))


if __name__ == "__main__":
    main()
